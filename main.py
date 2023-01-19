import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
import pandas as pd
import os


class Model:
    def __init__(self, regNum: str, previous_month_bal: float, total_fuel_purchased: float, begining_odo: int,
                 final_odo: int, last_odo: int):
        self.regNum = regNum
        self.previous_month_bal = previous_month_bal
        self.total_fuel_purchased = total_fuel_purchased
        self.begining_odo = begining_odo
        self.final_odo = final_odo
        self.last_odo = last_odo

    def search(self):
        os.chdir('/Users/dausdaneil/My Drive')
        df = pd.read_excel("KEND_JBU.xlsx", "senarai")
        bil = df["Bil"].values.tolist()
        brand = df["Jenama"].values.tolist()
        model = df["Model"].values.tolist()
        regNum = df["No Pendaftaran"].values.tolist()
        fuelType = df["Bahan Api"].values.tolist()
        ksl = df["KSL"].values.tolist()
        fullTank = df["Tangki Penuh"].values.tolist()
        status = df["Status"].values.tolist()

        z = 0
        while z != -1:
            if self.regNum.upper() == regNum[z] or self.regNum.upper()[3:] == regNum[z][4:]:
                return f"""{regNum[z]} is in the list.\n{brand[z]} {model[z]}\nKSL: {ksl[z]}\nFull Tank Capacity: {fullTank[z]}\nFuel Type: {fuelType[z]}\nStatus: {status[z]}"""
            elif len(self.regNum) == 0:
                return f"Registration Number Is Mandatory."
            else:
                z += 1

    def calculate_fuel_balance(self):
        os.chdir('/Users/dausdaneil/My Drive')

        jbu = load_workbook('KEND_JBU.xlsx')
        ws = jbu.active

        df = pd.read_excel("KEND_JBU.xlsx", "senarai")
        bil = df["Bil"].values.tolist()
        brand = df["Jenama"].values.tolist()
        model = df["Model"].values.tolist()
        regNum = df["No Pendaftaran"].values.tolist()
        fuelType = df["Bahan Api"].values.tolist()
        ksl = df["KSL"].values.tolist()
        fullTank = df["Tangki Penuh"].values.tolist()
        status = df["Status"].values.tolist()

        z = 0
        while z != -1:
            if self.regNum.upper() == regNum[z] or self.regNum.upper()[3:] == regNum[z][4:]:
                if self.last_odo < self.begining_odo:
                    fuel_bal = self.previous_month_bal - ((self.final_odo - self.last_odo) / (ksl[z]))
                else:
                    fuel_bal = fullTank[z] - ((self.final_odo - self.last_odo) / (ksl[z]))

                if fuelType[z] == 'PETROL':
                    total_purchased_cost = 2.05 * self.total_fuel_purchased
                else:
                    total_purchased_cost = 2.15 * self.total_fuel_purchased

                total_distance = self.final_odo - self.begining_odo

                total_fuel = self.total_fuel_purchased + self.previous_month_bal

                fuel_consumption = total_fuel - fuel_bal

                ksl = total_distance / fuel_consumption

                """
                response = messagebox.askyesno("Result",
                "\n(a). Previous Month Balance: " + str(self.previous_month_bal) +" Litre(s)" +
                "\n(b). Total Purchased: " + str(self.total_fuel_purchased) + " Litre(s)" +
                "\n(c). Total Fuel: " + "%.2f" % total_fuel + " Litre(s)" +
                "\n(d). Current Month Balance: " + "%.2f" % fuel_bal + " Litre(s)" +
                "\n(e). Fuel Consumption: " + "%.2f" % fuel_consumption + " Litre(s)" +
                "\nTotal Distance: " + str(total_distance) + " KM" +
                "\nKSL: " + "%.2f" % ksl +
                "\nTotal Costs: RM " + "%.2f" % total_purchased_cost +
                "\n\nDo you want to save this result for " + regNum[z])

                response = messagebox.askyesno("Result", "Fuel balance for " + regNum[z] + " is: " + "%.2f" % fuel_bal + "\nTotal Distance: " +\
                       str(total_distance) + "KM.\nTotal Fuel: " + "%.2f" % total_fuel + "\nTotal Costs: RM" +\
                       "%.2f" % total_purchased_cost + "\nFuel Consumption: " + "%.2f" % fuel_consumption +\
                       "\nKSL: " + "%.2f" % ksl + "\n\nDo you want to save this result for " + regNum[z])

                data = [self.total_fuel_purchased, float(total_purchased_cost), int(total_distance),
                        float("%.2f" % fuel_consumption), float("%.2f" % fuel_bal)]
                if response == 1:
                    for cell in range(10, 15):
                        ws.cell(row=bil[z] + 1, column=cell).value = data[cell - 10]
                    jbu.save('KEND_JBU.xlsx')

                """
                return "Fuel balance for " + regNum[z] + " is: " + "%.2f" % fuel_bal + "\nTotal Distance: " + \
                       str(total_distance) + "KM.\nTotal Fuel: " + "%.2f" % total_fuel + "\nTotal Costs: RM" + \
                       "%.2f" % total_purchased_cost + "\nFuel Consumption: " + "%.2f" % fuel_consumption + \
                       "\nKSL: " + "%.2f" % ksl

            else:
                z += 1


class View(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        # create widgets
        # regNum label
        self.label = ttk.Label(self, text='Registration Number:')
        self.label.grid(row=1, column=0)

        # regNum entry
        self.regNum_var = tk.StringVar()
        self.regNum_entry = ttk.Entry(self, textvariable=self.regNum_var, width=30)
        self.regNum_entry.grid(row=1, column=1, sticky=tk.NSEW)

        # previous_month_bal_label
        self.previous_month_bal_label = ttk.Label(self, text='Previous Balance:')
        self.previous_month_bal_label.grid(row=2, column=0)

        # previous_month_bal entry
        self.previous_month_bal_var = tk.StringVar()
        self.previous_month_bal_entry = ttk.Entry(self, textvariable=self.previous_month_bal_var, width=30)
        self.previous_month_bal_entry.grid(row=2, column=1, sticky=tk.NSEW)

        # total_purchased_label
        self.total_purchased_label = ttk.Label(self, text='Total Purchased:')
        self.total_purchased_label.grid(row=3, column=0)

        # total_purchased entry
        self.total_purchased_var = tk.StringVar()
        self.total_purchased_entry = ttk.Entry(self, textvariable=self.total_purchased_var, width=30)
        self.total_purchased_entry.grid(row=3, column=1, sticky=tk.NSEW)

        # bigining_odo label
        self.begining_odo_label = ttk.Label(self, text='Begining Odo:')
        self.begining_odo_label.grid(row=4, column=0)

        # bigining_odo entry
        self.begining_odo_var = tk.StringVar()
        self.begining_odo_entry = ttk.Entry(self, textvariable=self.begining_odo_var, width=30)
        self.begining_odo_entry.grid(row=4, column=1, sticky=tk.NSEW)

        # final_odo label
        self.final_odo_label = ttk.Label(self, text='Final Odo:')
        self.final_odo_label.grid(row=5, column=0)

        # final_odo entry
        self.final_odo_var = tk.StringVar()
        self.final_odo_entry = ttk.Entry(self, textvariable=self.final_odo_var, width=30)
        self.final_odo_entry.grid(row=5, column=1, sticky=tk.NSEW)

        # last_odo label
        self.last_odo_label = ttk.Label(self, text='Last Odo:')
        self.last_odo_label.grid(row=6, column=0)

        # last_odo entry
        self.last_odo_var = tk.StringVar()
        self.last_odo_entry = ttk.Entry(self, textvariable=self.last_odo_var, width=30)
        self.last_odo_entry.grid(row=6, column=1, sticky=tk.NSEW)

        # calculate button
        self.calculate_button = ttk.Button(self, text='Calculate', command=self.calculate_button_clicked)
        self.calculate_button.grid(row=2, column=3, padx=10)

        # search button
        self.search_button = ttk.Button(self, text='Search', command=self.search_button_clicked)
        self.search_button.grid(row=1, column=3, padx=10)

        # clear button
        self.clear_button = ttk.Button(self, text='Clear', command=self.clear_all_fields)
        self.clear_button.grid(row=3, column=3, padx=10)

        # message
        self.message_label = ttk.Label(self, text='', foreground='red')
        self.message_label.grid(row=7, column=1, sticky=tk.W)

        # set the controller
        self.controller = None

    def set_controller(self, controller):
        """
        Set the controller
        :param controller:
        :return:
        """
        self.controller = controller

    def search_button_clicked(self):
        """
        Handle button click event
        :return:
        """
        if self.controller:
            self.controller.search(self.regNum_var.get())

    def calculate_button_clicked(self):
        """
        Handle button click event
        :return:
        """
        try:
            if self.controller:
                self.controller.calculate(self.regNum_var.get(),
                                          float(self.previous_month_bal_var.get()),
                                          float(self.total_purchased_var.get()),
                                          int(self.begining_odo_var.get()),
                                          int(self.final_odo_var.get()),
                                          int(self.last_odo_var.get())
                                          )
        except ValueError:
            self.show_error("Fields cannot be empty.")

    def show_error(self, message):
        """
        Show an error message
        :param message:
        :return:
        """
        self.message_label['text'] = message
        self.message_label['foreground'] = 'red'
        self.message_label.after(3000, self.hide_message)
        self.regNum_entry['foreground'] = 'red'

    def show_success(self, message):
        """
        Show a success message
        :param message:
        :return:
        """
        self.message_label['text'] = message
        self.message_label['foreground'] = 'green'
        self.message_label.after(3000, self.hide_message)

        # reset the form
        self.regNum_entry['foreground'] = 'black'
        # self.regNum_var.set('')

    def hide_message(self):
        """
        Hide the message
        :return:
        """
        self.message_label['text'] = ''

    def clear_all_fields(self):
        """
        Clear all the fields
        :return:
        """
        self.regNum_var.set('')
        self.previous_month_bal_var.set('')
        self.total_purchased_var.set('')
        self.begining_odo_var.set('')
        self.final_odo_var.set('')
        self.last_odo_var.set('')


class Controller:
    def __init__(self, model, view):
        self.model = model
        self.view = view

    def search(self, regNum):
        try:
            self.model.regNum = regNum
            self.view.show_success(self.model.search())

        except IndexError:
            self.view.show_error(f"{regNum} not found.")

    def calculate(self, regNum, previous_month_bal, total_fuel_purchased, begining_odo, final_odo, last_odo):
        try:
            self.model.regNum = regNum
            self.model.previous_month_bal = previous_month_bal
            self.model.total_fuel_purchased = total_fuel_purchased
            self.model.begining_odo = begining_odo
            self.model.final_odo = final_odo
            self.model.last_odo = last_odo

            self.view.show_success(self.model.calculate_fuel_balance())

        except IndexError:
            self.view.show_error(f"{regNum} not found.")


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title('Reten Bulanan PKTN JBU')

        # create a model
        model = Model("", 0, 0, 0, 0, 0)

        # create a view and place it on the root window
        view = View(self)
        view.grid(row=0, column=0, padx=10, pady=10)

        # create a controller
        controller = Controller(model, view)

        # set the controller to view
        view.set_controller(controller)


if __name__ == '__main__':
    app = App()
    app.mainloop()